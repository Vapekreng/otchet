import openpyxl

book = openpyxl.Workbook()
book.save("1.xlsx")

wb = openpyxl.load_workbook(filename = '1.xlsx')
ws = wb['Sheet']
ws.cell(row=1, column=1).value = 11
f = open("1.txt", 'r', encoding='UTF-8')
info = []
new_row = 0
for line in f:
    new_row += 1
    line = line.split()
    result = dict()
    result['surname'] = line[6]
    ws.cell(row=new_row, column=1).value = result['surname']
    result['name'] = line[7]
    ws.cell(row=new_row, column=2).value = result['name']
    result['patronymic'] = line[8]
    ws.cell(row=new_row, column=3).value = result['patronymic']
    result['part 1'] = line[-4]
    ws.cell(row=new_row, column=4).value = result['part 1']
    result['part 2'] = line[-3]
    ws.cell(row=new_row, column=5).value = result['part 2']
    info.append(result)
wb.save("1.xlsx")


